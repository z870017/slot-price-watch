"""輸出層：前端 JSON、Excel 比價表、Phase 0 可行性數字。

Phase 0 的成敗就看 poc_summary() 算出來的那幾個數字：
三站到底有幾台機種是重疊的、價差多大。這決定這個專案值不值得做下去。
"""

import csv
import json
import os
from datetime import datetime, timezone

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DOCS_DIR = os.path.join(ROOT, "docs")
OUT_DIR = os.path.join(ROOT, "out")


def poc_summary(comparison: list, counts: dict, site_keys: list) -> dict:
    """Phase 0 要交出去的關鍵數字。"""
    multi = [c for c in comparison if c["site_count"] >= 2 and c["in_stock"]]
    all_three = [c for c in multi if c["site_count"] >= 3]
    spreads = [c["spread"] for c in multi if c["spread"] > 0]
    pcts = [c["spread_pct"] for c in multi if c["spread"] > 0]

    total_saving = sum(spreads)
    return {
        "total_products": sum(counts.values()),
        "per_site": counts,
        "total_models": len(comparison),
        "comparable_models": len(multi),
        "three_site_models": len(all_three),
        "models_with_spread": len(spreads),
        "avg_spread_yen": round(sum(spreads) / len(spreads)) if spreads else 0,
        "median_spread_yen": sorted(spreads)[len(spreads) // 2] if spreads else 0,
        "max_spread_yen": max(spreads) if spreads else 0,
        "avg_spread_pct": round(sum(pcts) / len(pcts), 1) if pcts else 0.0,
        "total_potential_saving_yen": total_saving,
        "verdict": _verdict(len(multi), sorted(spreads)[len(spreads) // 2] if spreads else 0),
    }


def _verdict(comparable: int, median_spread: int) -> str:
    """把數字翻譯成一句給客戶看的結論。"""
    if comparable == 0:
        return "三站商品幾乎沒有重疊，比價系統的價值有限，建議改做「單站降價監控」。"
    if comparable < 20:
        return f"只有 {comparable} 台機種可跨站比價，量偏少。建議先確認是否要擴充更多來源站。"
    if median_spread < 5000:
        return f"有 {comparable} 台可比價，但價差中位數僅 ¥{median_spread:,}，省下的金額有限。"
    return (
        f"有 {comparable} 台機種可跨站比價，價差中位數 ¥{median_spread:,} — "
        f"值得繼續做 Phase 1（自動排程＋降價通知）。"
    )


def write_frontend_json(comparison, summary, changes, site_meta, warnings, run_meta):
    os.makedirs(DOCS_DIR, exist_ok=True)
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "run": run_meta,
        "sites": site_meta,
        "summary": summary,
        "warnings": warnings,
        "changes": changes[:300],
        "items": comparison,
    }
    path = os.path.join(DOCS_DIR, "data.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
    return path


def write_csv(comparison, summary, changes, site_meta):
    """輸出給 Google 試算表吃的 CSV。

    Google Sheets 的 IMPORTDATA() 會定期自己去抓這個檔案的網址，
    所以只要 GitHub Actions 每天把新的 CSV commit 上去，試算表就會自己更新，
    完全不需要 Google 的 API 憑證或 service account。

    注意 IMPORTDATA 只能讀「公開網址」，所以 repo 必須是 public。
    """
    os.makedirs(DOCS_DIR, exist_ok=True)
    site_keys = [s["key"] for s in site_meta]
    site_names = {s["key"]: s["name"] for s in site_meta}
    paths = {}

    # --- 主表：比價 ---
    path = os.path.join(DOCS_DIR, "prices.csv")
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        header = ["機種名", "規格", "廠商", "站數", "有庫存", "最低價", "最高價", "價差", "價差%", "最便宜店家"]
        for key in site_keys:
            header += [f"{site_names[key]} 價格", f"{site_names[key]} 連結"]
        header.append("備註")
        w.writerow(header)

        for row in comparison:
            line = [
                row["name"], row["spec"] or "", row["maker"] or "", row["site_count"],
                "是" if row["in_stock"] else "否",
                row["min_price"], row["max_price"], row["spread"], row["spread_pct"],
                site_names.get(row["cheapest_site"], ""),
            ]
            for key in site_keys:
                entry = row["sites"].get(key)
                line += [entry["price"] if entry else "", entry["url"] if entry else ""]
            line.append("含偏離行情報價，價差已排除該筆" if row.get("has_outlier") else "")
            w.writerow(line)
    paths["prices"] = path

    # --- 變動表：降價/上架/售完 ---
    kind_label = {
        "price_down": "降價", "price_up": "漲價", "listed": "新上架",
        "delisted": "已下架", "sold_out": "已售完", "restocked": "補貨",
    }
    path = os.path.join(DOCS_DIR, "changes.csv")
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["類型", "商品名稱", "店家", "原價", "現價", "差額", "連結"])
        order = {"price_down": 0, "restocked": 1, "listed": 2, "price_up": 3, "sold_out": 4, "delisted": 5}
        for c in sorted(changes, key=lambda x: (order.get(x["kind"], 9),
                                                (x["new_price"] or 0) - (x["old_price"] or 0))):
            diff = ""
            if c["old_price"] and c["new_price"]:
                diff = c["new_price"] - c["old_price"]
            w.writerow([
                kind_label.get(c["kind"], c["kind"]), c["raw_name"],
                site_names.get(c["site"], c["site"]),
                c["old_price"] or "", c["new_price"] or "", diff, c["url"],
            ])
    paths["changes"] = path

    # --- 摘要表：直接可以貼給客戶看的 Phase 0 數字 ---
    path = os.path.join(DOCS_DIR, "summary.csv")
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["項目", "數值"])
        w.writerow(["最後更新", datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M")])
        w.writerow(["抓取商品總數", summary["total_products"]])
        for key in site_keys:
            w.writerow([f"　{site_names[key]}", summary["per_site"].get(key, 0)])
        w.writerow(["辨識機種數", summary["total_models"]])
        w.writerow(["可跨站比價機種數", summary["comparable_models"]])
        w.writerow(["三站都有的機種", summary["three_site_models"]])
        w.writerow(["價差中位數", summary["median_spread_yen"]])
        w.writerow(["價差平均", summary["avg_spread_yen"]])
        w.writerow(["最大價差", summary["max_spread_yen"]])
        w.writerow(["全買最便宜可省", summary["total_potential_saving_yen"]])
        w.writerow(["本次變動筆數", len(changes)])
        w.writerow(["結論", summary["verdict"]])
    paths["summary"] = path

    return paths


def write_excel(comparison, summary, site_meta, review_queue):
    """Phase 0 的主要交付物：一張可以直接寄給客戶看的比價表。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    os.makedirs(OUT_DIR, exist_ok=True)
    wb = Workbook()

    # --- Sheet 1: 比價表 ---
    ws = wb.active
    ws.title = "比價表"
    site_keys = [s["key"] for s in site_meta]
    headers = ["機種名", "規格", "廠商", "站數", "最低價", "最高價", "價差", "價差%", "最便宜"]
    for key in site_keys:
        headers.append(next(s["name"] for s in site_meta if s["key"] == key))
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2933")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    highlight = PatternFill("solid", fgColor="FFF3CD")
    for row in comparison:
        line = [
            row["name"], row["spec"] or "", row["maker"] or "", row["site_count"],
            row["min_price"], row["max_price"], row["spread"], row["spread_pct"],
            next((s["name"] for s in site_meta if s["key"] == row["cheapest_site"]), ""),
        ]
        for key in site_keys:
            entry = row["sites"].get(key)
            line.append(entry["price"] if entry else "")
        ws.append(line)
        if row["site_count"] >= 2 and row["spread"] > 0:
            for cell in ws[ws.max_row]:
                cell.fill = highlight

    widths = [46, 9, 14, 7, 12, 12, 12, 9, 16] + [14] * len(site_keys)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    for col in range(5, 8):
        for cell in ws[get_column_letter(col)][1:]:
            cell.number_format = "¥#,##0"

    # --- Sheet 2: PoC 結論 ---
    ws2 = wb.create_sheet("Phase0 結論")
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 60
    rows = [
        ("抓取商品總數", summary["total_products"]),
        ("－ 各站明細", ", ".join(f"{k}: {v}" for k, v in summary["per_site"].items())),
        ("辨識出的機種數", summary["total_models"]),
        ("可跨站比價機種數（2 站以上、有在庫）", summary["comparable_models"]),
        ("三站都有的機種數", summary["three_site_models"]),
        ("有價差的機種數", summary["models_with_spread"]),
        ("價差中位數", f"¥{summary['median_spread_yen']:,}"),
        ("價差平均", f"¥{summary['avg_spread_yen']:,}"),
        ("最大價差", f"¥{summary['max_spread_yen']:,}"),
        ("平均價差比例", f"{summary['avg_spread_pct']}%"),
        ("全部買最便宜可省", f"¥{summary['total_potential_saving_yen']:,}"),
        ("", ""),
        ("結論", summary["verdict"]),
    ]
    for label, value in rows:
        ws2.append([label, value])
    for cell in ws2["A"]:
        cell.font = Font(bold=True)
    ws2["B13"].alignment = Alignment(wrap_text=True, vertical="top")

    # --- Sheet 3: 待人工確認 ---
    ws3 = wb.create_sheet("待人工確認")
    ws3.append(["相似度", "未自動配對的原因", "疑似同一台的機種", "商品原始標題", "來源站", "網址"])
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for item in review_queue[:500]:
        ws3.append([item["score"], item.get("reason", ""), item["candidate_name"],
                    item["raw_name"], item["site"], item["url"]])
    for col, w in zip("ABCDEF", [10, 18, 40, 46, 14, 52]):
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = "A2"

    stamp = datetime.now().strftime("%Y%m%d")
    path = os.path.join(OUT_DIR, f"比價表_{stamp}.xlsx")
    wb.save(path)
    return path
